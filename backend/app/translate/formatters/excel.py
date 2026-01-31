"""
Excel 表格格式处理器
支持 .xlsx 文件的翻译，保持表格结构
"""
import uuid
import asyncio
from pathlib import Path
from typing import Callable, Optional
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

from . import BaseFormatter
from ...config import settings


class ExcelFormatter(BaseFormatter):
    """
    Excel 表格处理器

    使用 openpyxl 库处理 Excel 文件
    保持表格结构、公式等
    """

    def translate(
        self,
        source_path: str,
        target_lang: str,
        ai_translator,
        progress_callback: Optional[Callable[[int, int], None]] = None
    ) -> str:
        """
        翻译 Excel 表格（同步包装器，调用异步方法）

        Args:
            source_path: 源文件路径
            target_lang: 目标语言
            ai_translator: AI 翻译器实例
            progress_callback: 进度回调函数

        Returns:
            str: 翻译结果文件路径
        """
        # 检查翻译器是否支持并发方法
        if hasattr(ai_translator, 'translate_batch_async_concurrent'):
            # 获取线程数配置
            thread_count = getattr(ai_translator, 'thread_count', 5)

            # 创建新的事件循环并运行异步方法（不关闭循环，避免 AsyncOpenAI 客户端引用错误）
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            try:
                return loop.run_until_complete(
                    self.translate_async(source_path, target_lang, ai_translator, thread_count, progress_callback)
                )
            finally:
                # 不关闭循环，让 asyncio 自动管理
                asyncio.set_event_loop(None)
        else:
            # 使用原有同步逻辑
            # 加载工作簿
            wb = load_workbook(source_path)

            # 收集所有需要翻译的文本单元格
            cells_to_translate = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        # 只翻译字符串类型的单元格
                        if cell.value and isinstance(cell.value, str) and cell.value.strip():
                            # 跳过公式
                            if not cell.value.startswith('='):
                                cells_to_translate.append({
                                    'sheet': sheet,
                                    'cell': cell,
                                    'text': cell.value
                                })

            total_count = len(cells_to_translate)

            # 批量翻译
            batch_size = 20
            for i in range(0, total_count, batch_size):
                batch = cells_to_translate[i:i + batch_size]
                texts = [item['text'] for item in batch]

                # 翻译
                translated = ai_translator.translate_batch(texts, target_lang)

                # 更新单元格
                for j, item in enumerate(batch):
                    if j < len(translated):
                        item['cell'].value = translated[j]

                # 更新进度
                if progress_callback:
                    progress_callback(min(i + batch_size, total_count), total_count)

            # 保存结果
            result_path = self._generate_result_path(source_path)
            wb.save(result_path)
            wb.close()

            return result_path

    async def translate_async(
        self,
        source_path: str,
        target_lang: str,
        ai_translator,
        thread_count: int = 5,
        progress_callback: Optional[Callable[[int, int], None]] = None
    ) -> str:
        """
        异步翻译 Excel 表格（支持并发）

        Args:
            source_path: 源文件路径
            target_lang: 目标语言
            ai_translator: AI 翻译器实例
            thread_count: 并发线程数
            progress_callback: 进度回调函数

        Returns:
            str: 翻译结果文件路径
        """
        # 加载工作簿
        wb = load_workbook(source_path)

        # 收集所有需要翻译的文本单元格
        cells_to_translate = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    # 只翻译字符串类型的单元格
                    if cell.value and isinstance(cell.value, str) and cell.value.strip():
                        # 跳过公式
                        if not cell.value.startswith('='):
                            cells_to_translate.append({
                                'sheet': sheet,
                                'cell': cell,
                                'text': cell.value
                            })

        total_count = len(cells_to_translate)

        # 检查是否支持并发翻译
        if hasattr(ai_translator, 'translate_batch_async_concurrent'):
            # 使用并发翻译
            batch_size = 20  # 并发模式下可以增大批次
            for i in range(0, total_count, batch_size):
                batch = cells_to_translate[i:i + batch_size]
                texts = [item['text'] for item in batch]

                # 翻译
                translated = await ai_translator.translate_batch_async_concurrent(
                    texts=texts,
                    target_lang=target_lang,
                    max_concurrency=thread_count,
                    progress_callback=progress_callback
                )

                # 更新单元格
                for j, item in enumerate(batch):
                    if j < len(translated):
                        item['cell'].value = translated[j]
        else:
            # 降级到普通异步翻译
            for i, item in enumerate(cells_to_translate):
                translated = await ai_translator.translate_text_async(item['text'], target_lang)
                item['cell'].value = translated

                if progress_callback:
                    progress_callback(i + 1, total_count)

        # 保存结果
        result_path = self._generate_result_path(source_path)
        wb.save(result_path)
        wb.close()

        return result_path

    def _generate_result_path(self, source_path: str) -> str:
        """生成结果文件路径"""
        source = Path(source_path)
        filename = f"{source.stem}_translated_{uuid.uuid4().hex[:8]}{source.suffix}"
        return str(settings.TRANSLATE_DIR / filename)

    def extract_content(self, file_path: str, max_chars: int = 5000) -> dict:
        """
        提取 Excel 文件内容用于预览

        Args:
            file_path: 文件路径
            max_chars: 最大提取字符数

        Returns:
            dict: 包含 content 列表、total_chars、truncated、format
        """
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, data_only=True)

            content_list = []
            total_chars = 0
            truncated = False
            total_cells = 0

            for sheet_idx, sheet in enumerate(wb.worksheets):
                for row_idx, row in enumerate(sheet.iter_rows()):
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.strip():
                            total_cells += 1
                            text = cell.value.strip()

                            # 跳过公式
                            if text.startswith('='):
                                continue

                            if total_chars + len(text) > max_chars:
                                remaining = max_chars - total_chars
                                if remaining > 0:
                                    content_list.append({
                                        'type': 'cell',
                                        'text': text[:remaining],
                                        'index': total_cells,
                                        'location': f"{sheet.title} - 行{row_idx+1}"
                                    })
                                truncated = True
                                break

                            content_list.append({
                                'type': 'cell',
                                'text': text,
                                'index': total_cells,
                                'location': f"{sheet.title} - 行{row_idx+1}"
                            })
                            total_chars += len(text)

                    if truncated:
                        break
                if truncated:
                    break

            wb.close()

            return {
                'content': content_list,
                'total_chars': total_chars,
                'truncated': truncated,
                'format': 'xlsx',
                'total_sheets': len(wb.worksheets),
                'total_cells': total_cells
            }
        except Exception as e:
            return {
                'content': [],
                'total_chars': 0,
                'truncated': False,
                'format': 'xlsx',
                'error': str(e)
            }
